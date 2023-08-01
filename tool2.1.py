"This program imports the data from multiple time sheets into a master data file. All files are selected by the user. The program prints duplicate rows and organizes data by name and date"

import openpyxl
from openpyxl import Workbook, load_workbook
import time
import tkinter as tk
from tkinter import Tk, filedialog, messagebox, scrolledtext, Checkbutton, Label
from datetime import datetime, timedelta
import os
import sys
import re
import pandas as pd
import subprocess
from PIL import ImageTk, Image
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, PatternFill


## Notes ##
# Some features are built for the exe version and will not run in the python version correctly. These are the image display (since in the exe the image is packaged with the exe) and the initial directory for askopenfilename(s). Os.getcwd() is correct for the exe. Use os.path.dirname(__file__) for the python file

## General workflow ##
# The program begins by inviting the user to select an output file and then a list of timesheets to add their data to the output file
# It checks Master sheet validity making sure the file is not open and is usable this is done with the test open file method.
# It gets the name of the sheet it is inserting data to from get_sheet_name
# Then it begins to process the master file, first opening and converting all existing dates to strings so that they can be sorted later on (read_rows)
# Next it extracts the data from the master file including the cells formatting using extract_data_with_formatting
# Then it begins to process the timesheets one at a time calling the run_timesheet method. It first checks if each file is usable with test_open_file and then extracts all the data and writes it to self.rows_to_insert
# If there were any corrupted or open files it prints out those error messages
# During timesheet processing if missing data is found a red highlight is attached to that cell, otherwise the formatting is default
# Next, it combines the rows from the original master file and the new rows to insert into self.combined_rows
# Then the program removes empty rows (defined as a row without a name or date or hours) and sorts by name and date
# Finally the program writes these rows to the master file and master sheet. First deleting and clearing formatting of the sheet then inserting new rows
# Next the system performs data checks for duplicate rows ignoring columns past column K
# The program opens the master file once this is complete and asks the user if they would like to delete the duplicate rows (if any)

## ONE KEY LIMITATION IS IN THE TIMESHEET FILES, DATA ROWS WITH MORE THAN 3 EMPTY ROWS ABOVE IT WILL NOT BE INCLUDED IN THE DATA. MAKE SURE ALL TIMESHEETS HAVE ALL THEIR DATA SEPARATED BY ONE OR TWO EMPTY ROWS NOT MORE

# Custom exception for handling open file permission error, exits the program upon error so sheet is not tampered with
class FileOpenError(Exception):
	pass

class MyClass:
	# Class used so that file paths selected can be global
	def __init__(self):
		# declares global variables
		self.Master_file_path = None  # File path of output code
		self.duplicate_row_list = [] # list to have the option to delete cells after consulting master sheet
		self.Master_sheet_name = None #once sheet name is found, keep for whole run
		self.formatted_rows = [] #gets the old existing rows with formatting
		self.rows_to_insert = [] #gets the new rows
		self.combined_rows = [] #all combined rows
		self.found_gaps = False #flag for gaps found in new rows to display to user
	
	def run_timesheets(self):
		"This method operates first, allowing user to select files and then running through all the timesheets"
		# resets master varaibles to none before each new request
		self.duplicate_row_list = []
		self.formatted_rows = []
		self.rows_to_insert = []
		self.found_gaps = False
		
		# Configure buttons so user cannot repress buttons during a run
		timesheet_button.config(state = tk.DISABLED)
		instructions_button.config(state = tk.DISABLED)
		
		# Clear display at beginning of program
		scroll_text.configure(state='normal')
		scroll_text.delete('1.0', tk.END) 
		scroll_text.configure(state = 'disabled')
		
		# get output file path and timesheets, start selection where .py or .exe file is located
		self.Master_file_path = filedialog.askopenfilename(initialdir=os.getcwd(), title="Select Master sheet to write data to", filetypes=(("Master Sheet","*.xlsx"),))
		
		#check if no file selected and return to initial menu
		if not self.Master_file_path:
			scroll_text.configure(state='normal')
			scroll_text.insert(tk.END, "Please select an output file to run the program.\n") 
			scroll_text.configure(state = 'disabled')
			timesheet_button.config(state = tk.ACTIVE)
			instructions_button.config(state = tk.ACTIVE)
			return
		
		# Get timesheet file paths
		Timesheet_file_paths = filedialog.askopenfilenames(initialdir = os.getcwd(), title="Select Timesheets to insert", filetypes=(("Timesheets","*.xlsx"),))
		num_paths = len(Timesheet_file_paths)
		
		#check if no timesheets selected and return to intial menu
		if num_paths == 0:
			scroll_text.configure(state='normal')
			scroll_text.insert(tk.END, "Please select at least one timesheet to run the program.\n") 
			scroll_text.configure(state = 'disabled')
			timesheet_button.config(state = tk.ACTIVE)
			instructions_button.config(state = tk.ACTIVE)
			return
			
		# get name of sheet you are inserting data to from helper method
		self.Master_sheet_name = self.get_sheet_name(Timesheet_file_paths[0])
		
		# Test open master file path using helper method
		result,text = self.test_open_file(self.Master_file_path, self.Master_sheet_name)
		if not result:
			scroll_text.configure(state='normal')
			scroll_text.insert(tk.END, text) #prints the error message
			scroll_text.configure(state = 'disabled')
			timesheet_button.config(state = tk.ACTIVE)
			instructions_button.config(state = tk.ACTIVE)
			return
			
		#Initial Display
		scroll_text.configure(state='normal')
		scroll_text.insert(tk.END, "Accessing existing data...\n") 
		scroll_text.configure(state = 'disabled')
		root.update() #used to keep program from stalling throughout code
		
		# Convert dates from datetime to string for sorting
		rows = self.read_rows(self.Master_file_path)
		scroll_text.configure(state='normal')
		scroll_text.delete('1.0', tk.END) 
		scroll_text.insert(tk.END, "Accessing existing data... ...\n") 
		scroll_text.configure(state = 'disabled')
		root.update()
		
		# get the existing formatted rows from the sheet
		self.formatted_rows = self.extract_data_with_formatting(self.Master_file_path,self.Master_sheet_name)
		scroll_text.configure(state='normal')
		scroll_text.delete('1.0', tk.END) 
		scroll_text.insert(tk.END, "Accessing existing data... ... ...\n") 
		scroll_text.configure(state = 'disabled')
		root.update()
		
		def check_file_name(file_path):
			# Define the regular expression pattern to match "Approved" followed by two letters A-Z
			pattern = r"Approved [A-Z]{2}"

			# Extract the filename from the file path
			file_name = file_path.split("\\")[-1]

			# Search for the pattern in the file name
			match = re.search(pattern, file_name)

			# If a match is found, return True, otherwise return False
			return bool(match)
		
		# go through all the timesheet file paths and add to self.rows_to_insert, i and num_paths used for progress updates
		text_list = [] #used to store all error messages to print after run
		for i, timesheet in enumerate(Timesheet_file_paths):
			if check_file_name(timesheet):
				result, text = self.test_open_file(timesheet) #test open method and get error message 
				if result: #if sheet okay run timesheet
					self.run_timesheet(timesheet, i, num_paths)
				if not result: #else append error message to text_list to print
					text_list.append(text)
			else:
				text_list.append(f"\n[UNAPPROVED] File '{os.path.basename(timesheet)}' is not usable because it is not approved and doesn't match the naming convention.\n")
		
		# Display all error messages
		scroll_text.configure(state='normal')
		for message in text_list:
			scroll_text.insert(tk.END, message)
		scroll_text.insert(tk.END, "\n")
		scroll_text.configure(state = 'disabled')
		
		#Display configurations
		scroll_text.configure(state='normal')
		scroll_text.insert(tk.END, "\n\n\nData gathered, organizing ... \n") 
		scroll_text.configure(state = 'disabled')
		root.update()
		
		# Combine the existing and new data rows
		self.combined_rows = self.combine_data_with_formatting(self.formatted_rows, self.rows_to_insert)
		
		# Block to display updates cleanly while keeping error messages from above
		scroll_text.configure(state = 'normal')
		# Get existing minus last line to print updates
		cur_inp = scroll_text.get("1.0", tk.END) #get text on screen
		lines = cur_inp.split('\n')
		if lines[-3]:
			lines = lines[:-3]
		cur_inp_without_last_line = '\n'.join(lines) #keep up to last 3
		scroll_text.delete('1.0', tk.END) #clear then insert with new progress line
		scroll_text.insert(tk.END, cur_inp_without_last_line) 
		scroll_text.insert(tk.END, "Data gathered, organizing ... ... \n") 
		scroll_text.configure(state = 'disabled')
		root.update()
		
		# remove empty rows
		self.combined_rows = self.remove_rows_with_empty_values(self.combined_rows)
		
		scroll_text.configure(state = 'normal')
		cur_inp = scroll_text.get("1.0", tk.END)
		lines = cur_inp.split('\n')
		if lines[-3]:
			lines = lines[:-3]
		cur_inp_without_last_line = '\n'.join(lines)
		scroll_text.delete('1.0', tk.END) 
		scroll_text.insert(tk.END, cur_inp_without_last_line) 
		scroll_text.insert(tk.END, "Data gathered, organizing ... ... ...\n") 
		scroll_text.configure(state = 'disabled')
		root.update()
		
		# Sort the rows
		self.sort_combined_rows()
		
		scroll_text.configure(state = 'normal')
		cur_inp = scroll_text.get("1.0", tk.END)
		lines = cur_inp.split('\n')
		if lines[-3]:
			lines = lines[:-3]
		cur_inp_without_last_line = '\n'.join(lines)
		scroll_text.delete('1.0', tk.END) 
		scroll_text.insert(tk.END, cur_inp_without_last_line) 
		scroll_text.insert(tk.END, "Data gathered, organizing ... ... ... ...\n") 
		scroll_text.configure(state = 'disabled')
		root.update()
		
		# Delay for procesing
		time.sleep(1)
		
		#rewrite combined rows to Master sheet
		new_file_path = self.print_to_excel(self.combined_rows)
		
		# display configurations
		scroll_text.configure(state='normal')
		scroll_text.insert(tk.END, "Checking for duplicates ... \n") 
		scroll_text.configure(state = 'disabled')
		root.update()
		
		# delay for processing
		time.sleep(1)
		
		#check for duplicates
		self.check_duplicates(self.Master_file_path)
		if len(self.duplicate_row_list) == 0:
			scroll_text.configure(state='normal')
			scroll_text.insert(tk.END, "No duplicates found.\n") 
			scroll_text.configure(state = 'disabled')
		
		#delay for processing
		time.sleep(1)
		
		# Check for data gaps
		if self.found_gaps: #flag set during run timesheet method
			# Display configurations
			scroll_text.configure(state='normal')
			scroll_text.insert(tk.END, "\nMissing data identified, see red highlights in output.\n")
			scroll_text.see(tk.END)
			scroll_text.configure(state = 'disabled')
		else:
			# Display configurations
			scroll_text.configure(state='normal')
			scroll_text.insert(tk.END, "\nNo missing data identified.\n")
			scroll_text.see(tk.END)
			scroll_text.configure(state = 'disabled')
		
		# Display configurations
		scroll_text.configure(state='normal')
		scroll_text.insert(tk.END, "\nProcess Complete, opening output file.\n\n")
		scroll_text.see(tk.END)
		scroll_text.configure(state = 'disabled')
		
		# open output timesheet
		process = subprocess.Popen(f'explorer "{os.path.abspath(self.Master_file_path)}"')
		
		## Block to allow duplicate deletion if duplicate rows exist
		restart = True
		if len(self.duplicate_row_list) != 0: #if there is a duplicate
			while restart: #while loop used for if user does not close output file they can try deletion again
				duplicate_rows_text = ", ".join(str(item) for item in self.duplicate_row_list)
				result = messagebox.askquestion("Duplicates found.", f"Duplicate rows found: {duplicate_rows_text}. Would you like to delete them?")
				if result == "yes": #user clicked Yes on messagebox
					messagebox.showinfo("Attention", "Please close the output excel file before pressing OK.")
					check = self.delete_duplicates(self.Master_file_path) #delete duplicates, check is true false flag for master file open
					# Display configurations
					if check:
						scroll_text.configure(state='normal')
						scroll_text.insert(tk.END, f"Deletion Complete, you deleted rows: {duplicate_rows_text}.\nReopening output sheet.\n")
						scroll_text.see(tk.END)
						scroll_text.configure(state = 'disabled')
						restart = False #break loop
						process = subprocess.Popen(f'explorer "{os.path.abspath(self.Master_file_path)}"')
					else:
						#redo this while loop, prompt user to close file again
						messagebox.showinfo("Attention", "You must close the output file to delete duplicate rows.")
				else: #user pressed No, break loop
					restart = False
					
		# Reset buttons
		timesheet_button.config(state = tk.ACTIVE)
		instructions_button.config(state = tk.ACTIVE)
		
	def get_sheet_name(self, timesheet_path):
		"Helper Method to get the sheet name to extract formatted data"
		file_date_str = timesheet_path.split("_")[-1].split(" ")[1]
		file_date = datetime.strptime(file_date_str, "%m%d%Y")
		month_year = file_date.strftime("%B %Y")
		self.Master_sheet_name = month_year
		return month_year
	
	def test_open_file(self, file_path, sheet_name = None): 
		"this method takes in a file_name and tests to see if there are any issues with it"
		#param sheet_name defaults to None unless set, this method can be used for master and timesheet files
		text = "" #define string for error message
		# Load the workbook
		try:
			workbook = openpyxl.load_workbook(file_path)
			if sheet_name is not None and sheet_name not in workbook.sheetnames: #if name provided and does not exist, create it
				new_sheet = workbook.create_sheet(title = sheet_name)
				text = f"{sheet_name} not in the Output file path. It has been created."
			return True, text #good result
		except PermissionError:
			#file is open
			if sheet_name is not None:
				text = "The output file you are trying to use is currently open. Please close the file and try again.\n"
			else:
				text = "\n[OPEN FILE] " + os.path.basename(file_path) + " is currently open. Please close the file and try again.\n"
			return False, text #bad result
		except ValueError as e:
			#file cannot be read
			text = "\n[UNREADABLE] Unable to read workbook: " + os.path.basename(file_path) + ".\nThis file is unusable and must be entered manually.\n"
			return False, text #bad result
		except Exception as e:
			#generic exception
			text = f"\n[UNKNOWN ERROR] An error occurred while processing the file: {os.path.basename(file_path)}\nError details: {e}\nPlease check the file or enter data manually.\n"
			return False, text #bad result
	
	def read_rows(self, file_path):
		"This converts the dates to strings of the master file before doing anything else"
		# open the file and sheet
		wb = openpyxl.load_workbook(file_path)
		sheet = wb[self.Master_sheet_name]

		# Initialize array to store rows
		rows = []

		# Initialize an array to store the dates
		raw_dates = []

		# Go through rows starting from row 4
		for row_idx, row in enumerate(sheet.iter_rows(min_row=4, values_only=True), start =4):
			# Extract the date from column E (index 4)
			date = row[4]

			# Check if the date is not None and not blank
			if date is not None and not str(date).isspace():
				# Convert the date to a string in the desired format (MM/DD/YYYY)
				if isinstance(date, datetime):
					date_str = date.strftime("%m/%d/%Y")
				else:
					date_str = str(date)

				raw_dates.append(date_str)

			# Check if any cell in the row has a value (not None and not blank)
			if any(cell_value and not str(cell_value).isspace() for cell_value in row):
				if row[4] is not None and row[1] is not None and row[10] is not None:
					if not (str(row[4]).isspace() or str(row[1]).isspace() or str(row[10]).isspace()):
						# Replace the date value in column E with the corresponding string from raw_dates
						if len(raw_dates) > 0:
							row = list(row)
							row[4] = raw_dates.pop(0)
							rows.append(tuple(row))
								
								 # Write the modified row back to the sheet
						for col_idx, cell_value in enumerate(row, start=1):
							sheet.cell(row=row_idx, column=col_idx, value=cell_value)

		# Save the modified Excel file
		wb.save(file_path)
		return rows

	def extract_data_with_formatting(self, file_path, sheet_name):
		"This method extracts the existing data with all formatting and saves it in rows_with_formatting"
		# Load the Excel workbook
		workbook = load_workbook(file_path)

		# Select the specified sheet
		sheet = workbook[sheet_name]

		# Initialize the array to store rows
		rows_with_formatting = []

		# Loop through rows and store data with formatting
		for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row, values_only=False):
			row_data = []
			for cell in row:
				cell_value = cell.value
				# copy all existing formatting from the extracted data
				cell_formatting = {
					"font": cell.font.copy(),
					"fill": cell.fill.copy(),
					"border": cell.border.copy(),
					"alignment": cell.alignment.copy(),
					"number_format": cell.number_format,
				}
				row_data.append((cell_value, cell_formatting))
			rows_with_formatting.append(row_data)

		return rows_with_formatting #contains tuple with cell_value and cell_formatting
	
	def print_to_excel(self, rows_with_formatting):
		"Prints the rows with formatting to the master Excel sheet"
		# Open the master file and select working sheet
		wb = openpyxl.load_workbook(self.Master_file_path)
		sheet = wb[self.Master_sheet_name]
		
		# Clear existing data from rows 4 onwards including formatting
		for row_index in range(4, sheet.max_row + 1):
			for col_index in range(1, sheet.max_column + 1):
				sheet.cell(row=row_index, column=col_index).value = None
				# Clear formatting (highlight)
				sheet.cell(row=row_index, column=col_index).fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
				sheet.cell(row=row_index, column=col_index).font = Font()
				sheet.cell(row=row_index, column=col_index).alignment = Alignment()
				# Reset the border to the default style (thin grey line)
				default_border = Border(left=Side(style='thin', color='FFC0C0C0'),
										right=Side(style='thin', color='FFC0C0C0'),
										top=Side(style='thin', color='FFC0C0C0'),
										bottom=Side(style='thin', color='FFC0C0C0'))
				sheet.cell(row=row_index, column=col_index).border = default_border
				
		# Loop through rows and print data with formatting
		for row_idx, row in enumerate(rows_with_formatting, start=4):
			for col_idx, (cell_value, cell_formatting) in enumerate(row, start=1):
				# Write cell value to the new sheet
				sheet.cell(row=row_idx, column=col_idx, value=cell_value)

				if cell_formatting is not None:
					# Apply formatting to the cell if cell_formatting is not None
					new_cell = sheet.cell(row=row_idx, column=col_idx)
					for key, value in cell_formatting.items():
						setattr(new_cell, key, value)

		# Save the modified workbook
		wb.save(self.Master_file_path)
		
	def display_instructions(self):
		"This method prints instructions for operating this program to the user on the scrolled text window"
		scroll_text.configure(state='normal')
		scroll_text.delete('1.0', tk.END) #clear
		scroll_text.insert(tk.END, "How to use this program:\n\nBEFORE USING:\nEnsure all timesheets have no data lines 3 empty rows or more below the previous data line.\nData with more than 2 empty rows above it will not be entered.\n\nBegin by pressing the 'Select output file and timesheets' button.\nThis will bring up a window that allows you to select files.\n\nFirst select the output file where you want the data to be entered.\nPress the 'Open' button.\n\nNext it will bring up a new window where you will select all of the timesheets to include.\nPress the 'Open' button once you have selected all the timesheets you want to include.\n\nPlease only select timesheets to insert from the same month.\n\nThe program will now gather the data and organize it in your selected output file.\nIt will then identify duplicate rows (if any exist) and give you the option to delete them.\nIt will also look for data gaps and highlight them in red.\nIt will not highlight cells with missing data that already have highlights.\n\nContact James Schroeder at JWI if there are any issues.\n\nEnjoy!") 
		scroll_text.configure(state = 'disabled')
	
	def run_timesheet(self,file_path, i, num_paths):
		"This is the driving method that runs a singular timesheet collects all the data and writes it to the master excel sheet"
		# Clear display at beginning of program so progress updates in one line
		scroll_text.configure(state='normal')
		scroll_text.delete('1.0', tk.END) 
		scroll_text.configure(state = 'disabled')
		
		#open timesheet
		workbook = openpyxl.load_workbook(file_path)
		
		# Select the active sheet in timesheet file, not specified since timesheet file only has one sheet
		sheet = workbook.active

		# Read the week start date from cell G3
		sunday_start = sheet['G3'].value
		month_year = sunday_start.strftime("%B %Y") #month_year becomes sheet name in master file

		# Calculate the dates for each day of the week
		dates = []
		for day_offset in range(7):
			date = sunday_start + timedelta(days=day_offset)
			dates.append(date.strftime("%m/%d/%Y"))

		# Assign the dates to respective variables
		Sunday, Monday, Tuesday, Wednesday, Thursday, Friday, Saturday = dates

		# Read the name of the person
		name = sheet['C3'].value

		# Initialize a list to store the rows
		rows = []

		# Get the maximum row number
		max_row = sheet.max_row

		# Initialize a counter for empty work descriptions
		empty_work_desc_counter = 0

		# Iterate over the rows starting from row 13
		for row in range(13, max_row + 1):
			# Read the work description from column A
			work_description = sheet['A' + str(row)].value
			
			# Read the note from column B
			note = sheet['B' + str(row)].value
			
			# Get the code for the hours and convert to non coded language
			pay_type_code = sheet['D' + str(row)].value
			if pay_type_code == 'ST':
				time_type = "Regular Hours"
			elif pay_type_code == 'DT':
				time_type = "Double Time"
			elif pay_type_code == 'OT':
				time_type = "Overtime"
			else:
				time_type = pay_type_code #unknown code stays the same (typo)
			
			##THIS LOOP STOPS SEARCHING AFTER 3 EMPTY ROWS OF WORK_DESCRIPTION. THIS IS A KEY LIMITING FEATURE##
			if work_description is None:
				# Increment the counter for empty work descriptions
				empty_work_desc_counter += 1
				
				# Check if three consecutive empty work descriptions are encountered
				if empty_work_desc_counter >= 3:
					break
				else:
					continue
			
			# Reset the counter for empty work descriptions
			empty_work_desc_counter = 0
			
			# Read the hours for each day from Sunday to Saturday
			hours = []
			for day in range(5, 12):
				cell_value = sheet.cell(row=row, column=day).value
				hours.append(cell_value if cell_value is not None else 0)
			
			# Create a row array with the work description, note, and hours
			row_data = [work_description, note, time_type] + hours
			
			# Append the row array to the list of rows
			rows.append(row_data)

		# Close the workbook
		workbook.close()
		
		## At this point all timesheet data has been read from the sheet. Now program saves correct data and formatting to rows_to_insert ##
		
		# Define the file path for the output workbook
		new_file_path = self.Master_file_path

		# Load the existing workbook or create a new one if it doesn't exist
		new_workbook = load_workbook(new_file_path)
		
		#Extract the master sheet name from the file name
		file_date_str = file_path.split("_")[-1].split(" ")[1]
		file_date = datetime.strptime(file_date_str, "%m%d%Y")
		month_year = file_date.strftime("%B %Y")
		self.Master_sheet_name = month_year
		
		# Select the active sheet or create a new one if it doesn't exist
		if month_year in new_workbook.sheetnames:
			new_sheet = new_workbook[month_year]
		else:
			new_sheet = new_workbook.active
			new_sheet.title = month_year

		# Read the name of the person
		name = sheet['C3'].value

		# Extract approver initials 
		removed_last_5 = file_path[:-5]
		last_space_index = removed_last_5.rindex(" ")
		approver_initials = removed_last_5[last_space_index:]
		approver_initials = approver_initials[1:]
		
		# this block looks for names of approvers and connect with initials
		if "Data Validation" in new_workbook.sheetnames:
			initials_sheet = new_workbook["Data Validation"]
			#goes to data validation sheet in file
			approver_names = []
			
			approver_column = None
			for cell in initials_sheet[1]:
				if cell.value == "Approver":
					# find the header Approver
					approver_column = cell.column
					break
			
			if approver_column is not None:
				row_index = 1
				while initials_sheet.cell(row=row_index + 1, column=approver_column).value is not None:
					# go down through the list and add each name to the approver_names list
					row_index += 1
					cell_value = initials_sheet.cell(row=row_index, column=approver_column).value
					approver_names.append(cell_value)
			# break each approver name into initials and create matched array of initials
			approver_initials_list = []
			for Name in approver_names:
				name_parts = Name.split()
				initials = name_parts[0][0] + name_parts[-1][0]
				approver_initials_list.append(initials)
			# Check for a match and replace with full name by matched index
			if approver_initials in approver_initials_list:
				index = approver_initials_list.index(approver_initials)
				approver_initials = approver_names[index]

		# Iterate over the dates and rows
		for date in dates:
			# Iterate over the rows
			for row_data in rows:
				# Extract the work description, note, and hours
				work_description, note, time_type, hours = row_data[0], row_data[1], row_data[2], row_data[3:]
				
				# Get the index of the current date in the dates list
				date_index = dates.index(date)
				
				# Code for second third column data
				code = work_description.split("-")[0]
				col_2 = "ISTHA" + code + "-BR16"
				col_3 = "ISTHA Task " + code
				if code == "4":
					col_3 = col_3 + " (Professional SVS Eng)"
				elif code == "5":
					col_3 = col_3 + " (UR Supp Splicing)"
				elif code == "2C":
					col_3 = col_3 + " (Prof SVS PM)"
				elif code == "11":
					col_3 = col_3 + " (Watch/Protect)"
				elif code == "2C":
					col_3 = col_3 + " (Prof SVS PM)"
				
				# Get the hours worked for the current date
				hours_worked = hours[date_index]
				
				# Skip rows with zero hours
				if hours_worked == 0:
					continue
				
				# Mini method to find empty data
				def is_none_or_spaces(s):
					if s is None or str(s).isspace():
						self.found_gaps = True #mark gaps found as true for correct user display
						return True
					return False
				
				# Create a red fill for empty cells
				red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
				
				# Determine if the cell needs highlighting by if the cell is blank or just spaces
				name_formatting = {"fill": red_fill} if is_none_or_spaces(name) else None
				note_formatting = {"fill": red_fill} if is_none_or_spaces(note) else None
				work_description_formatting = {"fill": red_fill} if is_none_or_spaces(work_description) else None
				hours_worked_formatting = {"fill": red_fill} if is_none_or_spaces(hours_worked) else None
				
				weekNum = 11 + self.get_week_of_month(date) #for correct hours column formatting by weeks
				
					# Create a row to insert with the cell values and None for cell formatting
				row_to_insert = [
					("", None),
					(name, name_formatting),
					(col_2, None),
					(col_3, None),
					(date, None),
					(note, note_formatting),
					(work_description, work_description_formatting),
					(time_type, None),
					(approver_initials, None),
					("Approved", None),
					(hours_worked, hours_worked_formatting)
				]
				#This code used to correctly format the hours by week columns with blank data for correct formatting
				if weekNum == 12:
					row_to_insert.append((hours_worked, None))
				elif weekNum == 13:
					row_to_insert.append((None, None))
					row_to_insert.append((hours_worked, None))
				elif weekNum == 14:
					row_to_insert.append((None, None))
					row_to_insert.append((None, None))
					row_to_insert.append((hours_worked, None))
				elif weekNum == 15:
					row_to_insert.append((None, None))
					row_to_insert.append((None, None))
					row_to_insert.append((None, None))
					row_to_insert.append((hours_worked, None))
				elif weekNum == 16:
					row_to_insert.append((None, None))
					row_to_insert.append((None, None))
					row_to_insert.append((None, None))
					row_to_insert.append((None, None))
					row_to_insert.append((hours_worked, None))
				
				self.rows_to_insert.append(row_to_insert) #add the row to the master list of rows to insert
		
		# update the progress readout
		progress_percentage = (i+1)/num_paths*100
		progress_text = f"Processing Timesheets ... {i+1}/{num_paths} {progress_percentage:.1f}%\n"
		scroll_text.configure(state = 'normal')
		scroll_text.insert(tk.END, progress_text)
		scroll_text.see(tk.END) #scroll to bottom
		scroll_text.configure(state = 'disabled')
		root.update() #yield control to main loop so that program prints updates and doesn't stall

		# Close the master file sheet
		new_workbook.close()
		
	def combine_data_with_formatting(self, existing_data, new_data):
		"This method performs the operations to combine the data extracted from extract_data_with_formatting and run_timesheet with their respective highlights and other formatting"
		# Create a dictionary to store the formatting for each cell in the existing data
		formatting_dict = {}
		for row_idx, row_data in enumerate(existing_data):
			for col_idx, (cell_value, cell_formatting) in enumerate(row_data):
				formatting_dict[(row_idx + 1, col_idx + 1)] = cell_formatting

		# Initialize the array to store rows with formatting
		combined_rows_with_formatting = []

		# Combine the data from existing rows with formatting and new data
		for row_data in existing_data:
			name_cell_value = row_data[1][0]
			combined_row_data = []

			for col_idx, (cell_value, _) in enumerate(row_data):
				cell_formatting = formatting_dict.get((len(combined_rows_with_formatting) + 1, col_idx + 1), None)
				combined_row_data.append((cell_value, cell_formatting))

			combined_rows_with_formatting.append(combined_row_data)

		# Append the new data to the combined_rows_with_formatting
		for new_row_data in new_data:
			combined_rows_with_formatting.append(new_row_data)

		return combined_rows_with_formatting

	def remove_rows_with_empty_values(self, rows_with_formatting):
		"This method runs on combined_rows_with_formatting to remove rows without a name or date"
		# Define a filtering function to check for None values in name or date
		def filter_rows(row_data):
			name = row_data[1][0]
			date = row_data[4][0]
			return name is not None and date is not None

		# Filter out rows with NoneType values for name or date
		filtered_rows = [row for row in rows_with_formatting if filter_rows(row)]

		return filtered_rows
		
	def sort_combined_rows(self):
		"This method runs on combined_rows_with_formatting to sort the rows by name and date"
		# Define a custom sorting function
		def custom_sort_key(row_data):
			# Extract the name, date, and hours from the row data
			name = row_data[1][0]
			date = row_data[4][0]
			return (name, date)

		# Sort the combined_rows using the custom sorting function
		self.combined_rows.sort(key=custom_sort_key)

	def get_week_of_month(self,date):
		"this is a helper method that calculates the week of the month for the correct hours column formatting by week"
		isOneMinus = False # flag to subtract one if month starts on a saturday
		date_obj = datetime.strptime(date, "%m/%d/%Y") #convert to datetime object
		first_day = datetime(date_obj.year, date_obj.month, 1) #find the first day
		first_weekday = first_day.weekday() #then the weekday
		if first_weekday == 5: #check if saturday
			isOneMinus = True
		adjusted_date = date_obj.day + first_weekday - 1
		week_number = adjusted_date // 7 + 1 #calculates
		if isOneMinus:
			return week_number -1
		else:
			return week_number	

	def check_duplicates(self, file_path):
		"This method checks for duplicate rows and prints them out to the user. It deliberately selects inserted rows so as not to remove highlights"
		wb = openpyxl.load_workbook(file_path)
		sheet = wb[self.Master_sheet_name]
		
		# iter rows
		rows = list(sheet.iter_rows(min_row =4, values_only=True))
		
		#create dictionary of saved rows and list of duplicates
		saved_rows = {}
		duplicates = []
		first_dup = False #for formatting
		scroll_text.configure(state='normal')
		# go through rows
		for row_index, row in enumerate(rows, start=4):
			#only search for dups between col B and col K. This is so rows without the inputted weekly column hours aren't considered
			key = tuple(row[1:11]) 
			if key not in saved_rows: #store first occurance of everything
				saved_rows[key] = row_index
			else: #not first occurance means duplicate
				if not first_dup:
					scroll_text.insert(tk.END, "\n")
					first_dup = True
				duplicates.append(row)
				scroll_text.configure(state='normal')
				scroll_text.insert(tk.END, f"Row {row_index} is a duplicate of Row {saved_rows[key]}\n") #prints dup line and paired line
				scroll_text.configure(state='disabled')
				self.duplicate_row_list.append(row_index)
				
		#no duplicates found, formatting
		if not duplicates:
			scroll_text.insert(tk.END, "\n") 
		scroll_text.configure(state='disabled')
		
	def delete_duplicates(self, file_path):
		"This method opens an Excel file, deletes rows from self.duplicate_row_list, and moves other rows up. It returns True/False for if the master file was open"
		# Load the workbook
		try:
			wb = openpyxl.load_workbook(file_path)
			sheet = wb[self.Master_sheet_name]

			# Sort the duplicate_row_list in reverse order to delete rows from the end first
			self.duplicate_row_list.sort(reverse=True)

			# Delete rows in self.duplicate_row_list, no need to reformat since deletion doesn't empty rows but deletes them entirely
			for row_index in self.duplicate_row_list:
				sheet.delete_rows(row_index)
			
			# Save the modified workbook
			wb.save(file_path)
			return True
		except PermissionError:
			scroll_text.configure(state='normal')
			scroll_text.insert(tk.END, "Master output file is currently open. Please close the file and try again.\n") 
			scroll_text.configure(state = 'disabled')
			return False

def resource_path(relative_path):
	"This method makes it so that when this program is packaged as an exe it can find the image path and open it without needing the file"
	try:
		base_path = sys._MEIPASS
	except Exception:
		base_path = os.path.abspath(".")
	return os.path.join(base_path,relative_path)

# Create an instance of the class
my_instance = MyClass()

# Create the main Tkinter window
root = tk.Tk()
root.title("Timesheet Program")

# Create a button to trigger the PDF highlighting process
timesheet_button = tk.Button(root, text="Select output file and timesheets", command=my_instance.run_timesheets)
timesheet_button.grid(row =1, column =1, sticky = "w", padx=10, pady =5)

instructions_button = tk.Button(root, text = "Press to display instructions", command =my_instance.display_instructions)
instructions_button.grid(row =3, column =1, sticky = "w", padx=10,pady =5)

close_button = tk.Button(root, text="Close", command=root.quit)
close_button.grid(row=3, column=2, padx = 25, pady = 5)

# Create a scrolled text box
scroll_text = scrolledtext.ScrolledText(root, width=91, height=30, state='disabled')
scroll_text.grid(row =2, column =1, columnspan =2, sticky = "wens", padx =5, pady=5)

## Image Display, scaled from 3900x2517, scaled down by factor of 22
#THIS IMAGE DISPLAY ONLY WORKS WHEN THE FILE IS PACKAGED AS AN EXE, FOR NON EXE CHANGE IMAGE_FILE_PATH TO ACTUAL FILE PATH OF JWI LOGO
image_label = Label(root)
image_label.grid(row =2, column = 0, sticky = "w", padx = 10, pady =10)
image_file_path = resource_path("JWI Gray Logo.jpg") #call the method for the filepath
image=Image.open(image_file_path)
image=image.resize((177,114),Image.Resampling.LANCZOS)
photo = ImageTk.PhotoImage(image)
image_label.configure(image=photo)
image_label.image = photo

root.grid_rowconfigure(2, weight =1) #expands row 2 to fill screen
root.grid_columnconfigure(1, weight =1) #expands column 1 to fill screen

# Sets the window location on the console
root.geometry("+175+175")

# Run the Tkinter event loop
root.mainloop()		
		
		
		